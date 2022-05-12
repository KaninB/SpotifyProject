##################################################################################################################
####  Purpose: Collect missing values in data from the file "cleaned.csv"                                     ####
####  Author: Samantha Sharp                                                                                  ####
####  For IU CSCI-B365 Final Project                                                                          ####                                                                                             ####
##################################################################################################################
import spotipy
from spotipy.oauth2 import SpotifyClientCredentials
import sys
import openpyxl as Workbook
import configparser

# Reads the file and assigns the ids the Spotify's Client Credentials for obtaining the data
config = configparser.ConfigParser()
config.read('config.cfg')
client_id = config.get('SPOTIFY', 'CLIENT_ID')
client_secret = config.get('SPOTIFY', 'CLIENT_SECRET')
redirect_uri = config.get('SPOTIFY', 'REDIRECT_ID')

# Instantiates the Credentials
auth_manager = SpotifyClientCredentials(client_id=client_id, client_secret=client_secret)
sp = spotipy.Spotify(client_credentials_manager=auth_manager)

# Path where the data from cleaned.csv is obtain
# File has to be in excel format to be manipulated in this program
path = "C:\\Users\\sshar\\Downloads\\cleaned - Copy.xlsx" #** Needs to be changed for new user
workbook = Workbook.load_workbook(path) # Starting workbook
sheet = workbook.active

# Directory of Column names to be referenced easily
ColNames = {}
Current = 0
for col in sheet.iter_cols(1, sheet.max_column):
    ColNames[col[0].value] = Current
    Current += 1

# Goes through each row to obtain missing values
for row in range(2, sheet.max_row):
    track_id = "spotify:track:" + sheet.cell(row=row, column=ColNames['Song.ID'] + 1).value

    # Skips the track ids if false
    if len(track_id) == 36:
        audio_info = sp.audio_features([track_id])
        track_info = sp.track(track_id)
        artists = track_info['artists']

        # Since the data is two sets of data put together, some need different data than others
        # This if/else statement allows for only the needed data to be obtained
        # This statement is for the songs obtained in the spotify top 200 data
        if sheet.cell(row=row, column=ColNames['Highest.Charting.Position'] + 1).value != 'NA':
            sheet.cell(row=row, column=ColNames['mode'] + 1).value = audio_info[0]['mode'] # Adds mode value -- 1 is Major modality, 0 is minor
            sheet.cell(row=row, column=ColNames['instrumentalness'] + 1).value = audio_info[0]['instrumentalness'] # Adds instrumentalness
            sheet.cell(row=row, column=ColNames['time_signature'] + 1).value = audio_info[0]['time_signature'] # Adds time signature
            sheet.cell(row=row, column=ColNames['explicit'] + 1).value = track_info['explicit'] # Adds boolean whether the song is explicit or not
            genre = sheet.cell(row=row, column=ColNames['Genre'] + 1).value # gets the genre from the already obtained values
            sheet.cell(row=row, column=ColNames['Genre'] + 1).value = genre[1:-1] # Removes the brackets

            id = [] # List for all ids of all artists
            for i in range(len(artists)):
                id.append(artists[i]['id'])
            sheet.cell(row=row, column=ColNames['id_artists'] + 1).value = ', '.join(id) # Adds id list as a String

        # This is for the songs obtained from all the Spotify songs data
        else:
            followers_all = [] # List for all followers of the artists
            for i in range(len(artists)):
                artist_info = sp.artist(artists[i]['id']) # Artist information list
                followers = artist_info['followers'] # Retrieves a list of two values: href and follower amount
                followers_all.append(followers['total']) # Adds actual value for followers to list
            sheet.cell(row=row, column=ColNames['Artist.Followers'] + 1).value = max(followers_all) # Finds the most amount of followers of all the artists
            # ** This is done because the artist with the bigger follower will add new listeners because they are well known **

            genres_all = [] # List of genres
            for i in range(len(artists)):
                artist_id = artists[i]['id'] # ID of the artist
                artist = sp.artist(artist_id) # Data for the artist
                genres = artist['genres'] # Genres associated with the artist
                for j in range(len(genres)):
                    genre = genres[j] # Genre in the list of genres
                    if genre not in genres_all: # Prevents repeat values in list
                        genres_all.append(genre)
            if not genres_all : genres_all.append("NA") # No genre if the artist(s) has no genre associated with them
            sheet.cell(row=row, column=ColNames['Genre'] + 1).value = ', '.join(genres_all) # Adds genre list as a String

            key = sheet.cell(row=row, column=ColNames['key'] + 1).value # Uses the value given in Key column and converts it to Chord
            # Dictionary for transposing key to chord formatted Key : Chord
            dict = {0 : "C", 
                    1 : "C#/Db", 
                    2 : "D", 
                    3 : "D#/Eb", 
                    4 : "E", 
                    5 : "F", 
                    6 : "F#/Gb", 
                    7 : "G", 
                    8 : "G#/Ab", 
                    9 : "A", 
                    10 : "A#/Bb",
                    11 : "B"}
            
            sheet.cell(row=row, column=ColNames['Chord'] + 1).value = dict[key] # Adds cord to data

            artist = sheet.cell(row=row, column=ColNames['Artist'] + 1).value
            sheet.cell(row=row, column=ColNames['Artist'] + 1).value = artist[1:-1] # Removes brackets from Artists names

            artists_id = sheet.cell(row=row, column=ColNames['id_artists'] + 1).value
            sheet.cell(row=row, column=ColNames['id_artists'] + 1).value = artists_id[1:-1] # Removes brackets from artists ids

            # Converts explicit number to boolean value to make data uniform
            if (sheet.cell(row=row, column=ColNames['explicit'] + 1).value == 1):
                sheet.cell(row=row, column=ColNames['explicit'] + 1).value = 'TRUE'
            if (sheet.cell(row=row, column=ColNames['explicit'] + 1).value == 0):
                sheet.cell(row=row, column=ColNames['explicit'] + 1).value = 'FALSE'

# Saves workbook to be saved in excel
workbook.save(path)
